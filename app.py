from flask import Flask, render_template, request, redirect, url_for, flash, g, jsonify, send_file, make_response
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
import sqlite3
import bcrypt
import os
from datetime import date, datetime, timedelta
import random
from io import BytesIO
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

app = Flask(__name__)
app.secret_key = 'supersecretkey'
DB = 'crm.db'

login_manager = LoginManager(app)
login_manager.login_view = 'login'

class User(UserMixin):
    def __init__(self, id, username, role):
        self.id = id
        self.username = username
        self.role = role

@login_manager.user_loader
def load_user(user_id):
    db = get_db()
    c = db.cursor()
    c.execute("SELECT id, username, role FROM users WHERE id = ?", (user_id,))
    row = c.fetchone()
    return User(row[0], row[1], row[2]) if row else None

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB, check_same_thread=False)
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db is not None:
        db.close()

def init_db():
    if not hasattr(app, 'db_initialized'):
        db = sqlite3.connect(DB)
        c = db.cursor()

        # TABLES
        c.execute('''
        CREATE TABLE IF NOT EXISTS clients (
            id INTEGER PRIMARY KEY,
            business_type TEXT NOT NULL,
            business_name TEXT NOT NULL,
            contact_first TEXT,
            contact_last TEXT,
            phone TEXT,
            email TEXT,
            street TEXT,
            street2 TEXT,
            city TEXT,
            postcode TEXT,
            country TEXT,
            bacs_details TEXT,
            custom1 TEXT,
            custom2 TEXT,
            custom3 TEXT,
            default_interest_rate REAL DEFAULT 0.0
        )
        ''')

        c.execute('''
        CREATE TABLE IF NOT EXISTS cases (
            id INTEGER PRIMARY KEY,
            client_id INTEGER NOT NULL,
            debtor_business_type TEXT,
            debtor_business_name TEXT,
            debtor_first TEXT,
            debtor_last TEXT,
            phone TEXT,
            email TEXT,
            street TEXT,
            street2 TEXT,
            city TEXT,
            postcode TEXT,
            country TEXT,
            status TEXT DEFAULT 'Open',
            substatus TEXT,
            open_date TEXT DEFAULT (date('now')),
            custom1 TEXT,
            custom2 TEXT,
            custom3 TEXT,
            interest_rate REAL,
            FOREIGN KEY (client_id) REFERENCES clients(id) ON DELETE CASCADE
        )
        ''')

        c.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            password_hash BLOB NOT NULL,
            role TEXT DEFAULT 'user'
        )
        ''')

        c.execute('''
        CREATE TABLE IF NOT EXISTS money (
            id INTEGER PRIMARY KEY,
            case_id INTEGER NOT NULL,
            type TEXT NOT NULL,
            amount REAL NOT NULL,
            transaction_date TEXT DEFAULT (date('now')),
            created_by INTEGER NOT NULL,
            note TEXT,
            FOREIGN KEY (case_id) REFERENCES cases(id) ON DELETE CASCADE,
            FOREIGN KEY (created_by) REFERENCES users(id)
        )
        ''')

        c.execute('''
        CREATE TABLE IF NOT EXISTS notes (
            id INTEGER PRIMARY KEY,
            case_id INTEGER NOT NULL,
            type TEXT NOT NULL,
            created_by INTEGER NOT NULL,
            note TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (case_id) REFERENCES cases(id) ON DELETE CASCADE,
            FOREIGN KEY (created_by) REFERENCES users(id)
        )
        ''')

        # ADMIN
        c.execute("SELECT COUNT(*) FROM users")
        if c.fetchone()[0] == 0:
            hashed = bcrypt.hashpw(b'admin', bcrypt.gensalt())
            c.execute("INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)", ('admin', hashed, 'admin'))

        # DUMMY DATA
        c.execute("SELECT COUNT(*) FROM clients")
        if c.fetchone()[0] == 0:
            clients = [
                ("Limited", "Acme Corp", "John", "Doe", "01234 567890", "john@acme.com", "8.5"),
                ("Sole Trader", "Bob's Plumbing", "Bob", "Smith", "07700 900123", "bob@plumb.co.uk", "7.0"),
                ("Partnership", "Green & Co", "Sarah", "Green", "020 7946 0001", "sarah@green.co", "6.5"),
                ("Individual", "Freelance Designs", "Alex", "Taylor", "07890 123456", "alex@design.com", "9.0"),
                ("Limited", "Tech Solutions Ltd", "Mike", "Brown", "0113 496 0002", "mike@techsol.co.uk", "8.0")
            ]
            client_ids = []
            for cl in clients:
                c.execute('''
                    INSERT INTO clients (business_type, business_name, contact_first, contact_last, phone, email, default_interest_rate)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', cl)
                client_ids.append(c.lastrowid)

            debtor_types = ["Individual", "Sole Trader", "Limited", "Partnership"]
            statuses = ["Open", "On Hold", "Closed"]
            for client_id in client_ids:
                for i in range(5):
                    debtor_type = random.choice(debtor_types)
                    first = random.choice(["Emma", "James", "Olivia", "Liam", "Noah", "Ava"])
                    last = random.choice(["Wilson", "Davis", "Martinez", "Lee", "Clark", "Walker"])
                    business = f"{first} {last} Ltd" if debtor_type in ["Limited", "Partnership"] else None
                    c.execute('''
                        INSERT INTO cases 
                        (client_id, debtor_business_type, debtor_business_name, debtor_first, debtor_last, phone, email, postcode, status, substatus, interest_rate)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        client_id, debtor_type, business, first, last,
                        f"07{random.randint(100,999)} {random.randint(100000,999999)}",
                        f"{first.lower()}.{last.lower()}@example.com",
                        f"SW1A {random.randint(1,9)}AA",
                        random.choice(statuses),
                        random.choice(["Awaiting Docs", "In Court", "Payment Plan", None]),
                        float(clients[client_ids.index(client_id)][6])
                    ))
                    case_id = c.lastrowid

                    for _ in range(random.randint(3,7)):
                        typ = random.choice(["Invoice", "Payment", "Charge", "Interest"])
                        amount = round(random.uniform(50, 1500), 2)
                        days_ago = random.randint(1, 180)
                        tx_date = (datetime.now() - timedelta(days=days_ago)).strftime('%Y-%m-%d')
                        c.execute('''
                            INSERT INTO money (case_id, type, amount, created_by, transaction_date, note)
                            VALUES (?, ?, ?, ?, ?, ?)
                        ''', (case_id, typ, amount, 1, tx_date, f"{typ} entry" if random.random() > 0.5 else None))

                    for _ in range(random.randint(2,5)):
                        note_type = random.choice(["General", "Inbound Call", "Outbound Call", "Dispute"])
                        note_text = random.choice([
                            "Customer called to discuss balance",
                            "Sent reminder letter",
                            "Payment plan agreed",
                            "Dispute raised – awaiting proof",
                            "Left voicemail"
                        ])
                        note_date = (datetime.now() - timedelta(days=random.randint(1, 120))).strftime('%Y-%m-%d %H:%M:%S')
                        c.execute('''
                            INSERT INTO notes (case_id, type, created_by, note, created_at)
                            VALUES (?, ?, ?, ?, ?)
                        ''', (case_id, note_type, 1, note_text, note_date))

        db.commit()
        db.close()
        app.db_initialized = True

@app.before_request
def before_request():
    init_db()

# === FORMS ===
@app.route('/add_client', methods=['POST'])
@login_required
def add_client():
    data = request.form
    db = get_db()
    c = db.cursor()
    c.execute('''
        INSERT INTO clients 
        (business_type, business_name, contact_first, contact_last, phone, email, street, street2, city, postcode, country, bacs_details, custom1, custom2, custom3, default_interest_rate)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        data['business_type'], data['business_name'], data.get('contact_first'), data.get('contact_last'),
        data.get('phone'), data.get('email'), data.get('street'), data.get('street2'),
        data.get('city'), data.get('postcode'), data.get('country'), data.get('bacs_details'),
        data.get('custom1'), data.get('custom2'), data.get('custom3'), float(data.get('default_interest_rate', 0))
    ))
    db.commit()
    return redirect(url_for('dashboard'))

@app.route('/add_case', methods=['POST'])
@login_required
def add_case():
    data = request.form
    db = get_db()
    c = db.cursor()
    client_id = int(data['client_id'])
    c.execute("SELECT default_interest_rate FROM clients WHERE id = ?", (client_id,))
    client = c.fetchone()
    interest_rate = client['default_interest_rate'] if client else 0.0
    c.execute('''
        INSERT INTO cases 
        (client_id, debtor_business_type, debtor_business_name, debtor_first, debtor_last, phone, email, street, street2, city, postcode, country, status, substatus, custom1, custom2, custom3, interest_rate)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        client_id, data['debtor_business_type'], data.get('debtor_business_name'),
        data.get('debtor_first'), data.get('debtor_last'), data.get('phone'), data.get('email'),
        data.get('street'), data.get('street2'), data.get('city ⊆'), data.get('postcode'), data.get('country'),
        data.get('status', 'Open'), data.get('substatus'), data.get('custom1'), data.get('custom2'), data.get('custom3'), interest_rate
    ))
    db.commit()
    return redirect(url_for('dashboard'))

@app.route('/add_transaction', methods=['POST'])
@login_required
def add_transaction():
    data = request.form
    db = get_db()
    c = db.cursor()
    c.execute('''
        INSERT INTO money (case_id, type, amount, created_by, note)
        VALUES (?, ?, ?, ?, ?)
    ''', (data['case_id'], data['type'], float(data['amount']), current_user.id, data.get('note')))
    db.commit()
    return redirect(url_for('dashboard', case_id=data['case_id']))

@app.route('/add_note', methods=['POST'])
@login_required
def add_note():
    data = request.form
    db = get_db()
    c = db.cursor()
    c.execute('''
        INSERT INTO notes (case_id, type, created_by, note)
        VALUES (?, ?, ?, ?)
    ''', (data['case_id'], data['type'], current_user.id, data['note']))
    db.commit()
    return redirect(url_for('dashboard', case_id=data['case_id']))

# === SEARCH ===
@app.route('/search')
@login_required
def search():
    query = request.args.get('q', '').strip()
    field = request.args.get('field', 'debtor_first').strip()
    mode = request.args.get('mode', 'contains')

    if not query:
        return jsonify([])

    db = get_db()
    c = db.cursor()

    field_map = {
        'debtor_name': "(s.debtor_first || ' ' || s.debtor_last)",
        'postcode': "s.postcode",
        'email': "s.email",
        'phone': "s.phone",
        'client_name': "c.business_name",
        'client_code': "c.id"
    }
    col = field_map.get(field, "s.debtor_first")

    operator = '=' if mode == 'is' else 'LIKE'
    param = query if mode == 'is' else f'%{query}%'

    sql = f"""
        SELECT s.id as case_id, c.business_name, c.id as client_code,
               s.debtor_first, s.debtor_last, s.debtor_business_name,
               s.postcode, s.email, s.phone
        FROM cases s
        JOIN clients c ON s.client_id = c.id
        WHERE {col} {operator} ?
        ORDER BY c.business_name, s.id
        LIMIT 50
    """
    c.execute(sql, (param,))
    results = c.fetchall()
    db.close

    return jsonify([{
        'case_id': r['case_id'],
        'client': r['business_name'],
        'client_code': r['client_code'],
        'debtor': r['debtor_business_name'] or f"{r['debtor_first']} {r['debtor_last']}",
        'postcode': r['postcode'] or '',
        'email': r['email'] or '',
        'phone': r['phone'] or ''
    } for r in results])

# === REPORTS ===
@app.route('/report')
@login_required
def report():
    client_code = request.args.get('client_code', '').strip()
    db = get_db()
    c = db.cursor()

    if not client_code:
        return render_template('report.html', report_html='', client_code='')

    c.execute("SELECT id, business_name FROM clients WHERE id = ?", (client_code,))
    client = c.fetchone()
    if not client:
        return render_template('report.html', report_html='<p>Client not found.</p>', client_code=client_code)

    c.execute("""
        SELECT s.id as case_id, s.debtor_business_name, s.debtor_first, s.debtor_last,
               m.type, m.amount
        FROM cases s
        LEFT JOIN money m ON s.id = m.case_id
        WHERE s.client_id = ?
        ORDER BY s.id, m.transaction_date
    """, (client_code,))
    rows = c.fetchall()

    cases = {}
    for r in rows:
        case_id = r['case_id']
        if case_id not in cases:
            debtor = r['debtor_business_name'] or f"{r['debtor_first']} {r['debtor_last']}"
            cases[case_id] = {'debtor': debtor, 'Invoice': 0, 'Payment': 0, 'Charge': 0, 'Interest': 0}
        if r['type']:
            cases[case_id][r['type']] += r['amount']

    html = f"""
    <h2>Client Report: {client['business_name']} (ID: {client['id']})</h2>
    <table border="1" style="width:100%; border-collapse:collapse; font-family:Arial; font-size:14px;">
        <tr style="background:#f5f5f5;">
            <th>Case ID</th><th>Debtor</th><th>Invoice</th><th>Payment</th><th>Charge</th><th>Interest</th><th>Balance</th>
        </tr>
    """
    grand = {'Invoice': 0, 'Payment': 0, 'Charge': 0, 'Interest': 0}
    for case_id, data in cases.items():
        balance = data['Invoice'] + data['Charge'] + data['Interest'] - data['Payment']
        html += f"""
        <tr>
            <td>{case_id}</td>
            <td>{data['debtor']}</td>
            <td>£{data['Invoice']:,.2f}</td>
            <td>£{data['Payment']:,.2f}</td>
            <td>£{data['Charge']:,.2f}</td>
            <td>£{data['Interest']:,.2f}</td>
            <td>£{balance:,.2f}</td>
        </tr>
        """
        for t in ['Invoice', 'Payment', 'Charge', 'Interest']:
            grand[t] += data[t]

    grand_balance = grand['Invoice'] + grand['Charge'] + grand['Interest'] - grand['Payment']
    html += f"""
        <tr style="font-weight:bold; background:#eef;">
            <td colspan="2">TOTALS</td>
            <td>£{grand['Invoice']:,.2f}</td>
            <td>£{grand['Payment']:,.2f}</td>
            <td>£{grand['Charge']:,.2f}</td>
            <td>£{grand['Interest']:,.2f}</td>
            <td>£{grand_balance:,.2f}</td>
        </tr>
    </table>
    """
    return render_template('report.html', report_html=html, client_code=client_code)

@app.route('/export_excel')
@login_required
def export_excel():
    client_code = request.args.get('client_code')
    db = get_db()
    c = db.cursor()
    c.execute("SELECT id, business_name FROM clients WHERE id = ?", (client_code,))
    client = c.fetchone()
    if not client:
        return "Client not found", 404

    c.execute("""
        SELECT s.id as case_id, s.debtor_business_name, s.debtor_first, s.debtor_last,
               m.type, m.amount
        FROM cases s
        LEFT JOIN money m ON s.id = m.case_id
        WHERE s.client_id = ?
    """, (client_code,))
    rows = c.fetchall()

    cases = {}
    for r in rows:
        case_id = r['case_id']
        if case_id not in cases:
            debtor = r['debtor_business_name'] or f"{r['debtor_first']} {r['debtor_last']}"
            cases[case_id] = {'debtor': debtor, 'Invoice': 0, 'Payment': 0, 'Charge': 0, 'Interest': 0}
        if r['type']:
            cases[case_id][r['type']] += r['amount']

    wb = Workbook()
    ws = wb.active
    ws.title = "Client Report"

    # Header
    headers = ['Case ID', 'Debtor', 'Invoice', 'Payment', 'Charge', 'Interest', 'Balance']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Data + Totals
    grand = {'Invoice': 0, 'Payment': 0, 'Charge': 0, 'Interest': 0}
    for case_id, d in cases.items():
        balance = d['Invoice'] + d['Charge'] + d['Interest'] - d['Payment']
        row = [case_id, d['debtor'], d['Invoice'], d['Payment'], d['Charge'], d['Interest'], balance]
        ws.append(row)
        for t in ['Invoice', 'Payment', 'Charge', 'Interest']:
            grand[t] += d[t]

    # Grand Total
    grand_balance = grand['Invoice'] + grand['Charge'] + grand['Interest'] - grand['Payment']
    total_row = ['TOTALS', '', grand['Invoice'], grand['Payment'], grand['Charge'], grand['Interest'], grand_balance]
    ws.append(total_row)
    total_row_cells = ws[ws.max_row]
    for cell in total_row_cells:
        cell.font = Font(bold=True)

    # Format numbers with commas
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=7):
        for cell in row:
            cell.number_format = '#,##0.00'

    # Borders
    thin = Side(border_style="thin")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f"report_client_{client_code}.xlsx", as_attachment=True)

@app.route('/export_pdf')
@login_required
def export_pdf():
    client_code = request.args.get('client_code')
    db = get_db()
    c = db.cursor()
    c.execute("SELECT id, business_name FROM clients WHERE id = ?", (client_code,))
    client = c.fetchone()
    if not client:
        return "Client not found", 404

    c.execute("""
        SELECT s.id as case_id, s.debtor_business_name, s.debtor_first, s.debtor_last,
               m.type, m.amount
        FROM cases s
        LEFT JOIN money m ON s.id = m.case_id
        WHERE s.client_id = ?
    """, (client_code,))
    rows = c.fetchall()

    cases = {}
    for r in rows:
        case_id = r['case_id']
        if case_id not in cases:
            debtor = r['debtor_business_name'] or f"{r['debtor_first']} {r['debtor_last']}"
            cases[case_id] = {'debtor': debtor, 'Invoice': 0, 'Payment': 0, 'Charge': 0, 'Interest': 0}
        if r['type']:
            cases[case_id][r['type']] += r['amount']

    html = f"""
    <h1>Client Report: {client['business_name']} (ID: {client['id']})</h1>
    <table border="1" style="width:100%; border-collapse:collapse; font-family:Arial; font-size:12px;">
        <tr style="background:#ddd;">
            <th>Case ID</th><th>Debtor</th><th>Invoice</th><th>Payment</th><th>Charge</th><th>Interest</th><th>Balance</th>
        </tr>
    """
    for case_id, d in cases.items():
        balance = d['Invoice'] + d['Charge'] + d['Interest'] - d['Payment']
        html += f"""
        <tr>
            <td>{case_id}</td>
            <td>{d['debtor']}</td>
            <td>£{d['Invoice']:,.2f}</td>
            <td>£{d['Payment']:,.2f}</td>
            <td>£{d['Charge']:,.2f}</td>
            <td>£{d['Interest']:,.2f}</td>
            <td>£{balance:,.2f}</td>
        </tr>
        """
    grand = {t: sum(c[t] for c in cases.values()) for t in ['Invoice','Payment','Charge','Interest']}
    grand_balance = grand['Invoice'] + grand['Charge'] + grand['Interest'] - grand['Payment']
    html += f"""
        <tr style="font-weight:bold; background:#eee;">
            <td colspan="2">TOTALS</td>
            <td>£{grand['Invoice']:,.2f}</td>
            <td>£{grand['Payment']:,.2f}</td>
            <td>£{grand['Charge']:,.2f}</td>
            <td>£{grand['Interest']:,.2f}</td>
            <td>£{grand_balance:,.2f}</td>
        </tr>
    </table>
    """

    pdf = HTML(string=html).write_pdf()
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=report_client_{client_code}.pdf'
    return response

# === DASHBOARD ===
@app.route('/')
@app.route('/dashboard')
@login_required
def dashboard():
    db = get_db()
    c = db.cursor()

    c.execute("SELECT id, business_name FROM clients ORDER BY business_name")
    clients = c.fetchall()

    c.execute("""
        SELECT c.id as client_id, c.business_name, s.id as case_id, 
               s.debtor_business_name, s.debtor_first, s.debtor_last
        FROM clients c
        LEFT JOIN cases s ON c.id = s.client_id
        ORDER BY c.business_name, s.id
    """)
    all_cases = c.fetchall()

    selected_case = None
    case_client = None
    client_cases = []
    notes = []
    transactions = []
    balance = 0.0
    totals = {'Invoice': 0, 'Payment': 0, 'Charge': 0, 'Interest': 0}

    case_id = request.args.get('case_id')
    if case_id:
        c.execute("SELECT * FROM cases WHERE id = ?", (case_id,))
        selected_case = c.fetchone()
        if selected_case:
            c.execute("SELECT * FROM clients WHERE id = ?", (selected_case['client_id'],))
            case_client = c.fetchone()

            c.execute("SELECT id, debtor_business_name, debtor_first, debtor_last FROM cases WHERE client_id = ? ORDER BY id", (selected_case['client_id'],))
            client_cases = c.fetchall()

            c.execute('SELECT n.*, u.username FROM notes n JOIN users u ON n.created_by = u.id WHERE n.case_id = ? ORDER BY n.created_at DESC', (case_id,))
            notes = c.fetchall()

            c.execute('SELECT m.*, u.username FROM money m JOIN users u ON m.created_by = u.id WHERE m.case_id = ? ORDER BY m.transaction_date DESC, m.id DESC', (case_id,))
            transactions = c.fetchall()

            for t in transactions:
                amt = t['amount']
                typ = t['type']
                totals[typ] += amt
                if typ == 'Payment':
                    balance -= amt
                else:
                    balance += amt

    # Format balance with commas
    balance_str = f"£{balance:,.2f}"

    return render_template('dashboard.html',
                           clients=clients,
                           all_cases=all_cases,
                           selected_case=selected_case,
                           case_client=case_client,
                           client_cases=client_cases,
                           notes=notes,
                           transactions=transactions,
                           balance=balance_str,
                           totals=totals)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password'].encode('utf-8')
        db = get_db()
        c = db.cursor()
        c.execute("SELECT * FROM users WHERE username = ?", (username,))
        user = c.fetchone()
        if user and bcrypt.checkpw(password, user['password_hash']):
            login_user(User(user['id'], user['username'], user['role']))
            return redirect(url_for('dashboard'))
        flash('Invalid credentials')
    return render_template('login.html')

if __name__ == '__main__':
    app.run(debug=True)
